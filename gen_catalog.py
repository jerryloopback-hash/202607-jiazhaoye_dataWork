# -*- coding: utf-8 -*-
"""根据 _parsed_tables.json 生成《训练_数据源清单》，格式对齐模板 xlsx。
- Sheet1「所有数据表及数据字段」：数据表名称/数据表中文名称/字段名称/字段中文称/字段类型/截图/别名
- Sheet2「所有数据表」：表名/中文表名/引擎/自增/数据长度
有表注释的直接用注释；无注释的表/字段根据表名、字段名等推测中文名。
"""
import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

tables = json.load(open('_parsed_tables.json', encoding='utf-8'))

# ============ 无表注释的表 —— 人工推测中文表名 ============
# 依据前缀语义: c_ 商品/进销存, h_ 卡, j_ 文件回盘, l_ 平台&商户&权限,
#               m_ 交易/支付/分润(含 Ping++、银联), sys_ 系统, t_ 培训核心业务
TABLE_CN = {
    'c_goods_access_storage_code':            '出入库单据编号表',
    'c_goods_stock_record':                   '商品库存变动流水表',
    'h_card':                                 '实体卡表',
    'h_member_course_refund':                 '会员课程退款申请表',
    'j_upload_file_return':                   '上传文件回盘记录表',
    'l_admin':                                '平台管理员表',
    'l_merchant_settle_account':              '商户结算收款账户表',
    'l_permission':                           '权限资源表',
    'm_pingxx_coupon':                        'Ping++优惠券领用记录表',
    'm_pingxx_coupon_template':               'Ping++优惠券模板表',
    'm_pingxx_notify':                        'Ping++异步通知记录表',
    'm_pingxx_order':                         'Ping++支付订单表',
    'm_pingxx_royalty':                       'Ping++分润记录表',
    'm_pingxx_royalty_settlement':            'Ping++分润结算表',
    'm_pingxx_royalty_settlement_order_num':  'Ping++分润结算-订单号关联表',
    'm_pingxx_royalty_template':              'Ping++分润模板表',
    'm_pingxx_royalty_transaction':           'Ping++分润结算交易表',
    'm_pingxx_subapp':                        'Ping++子商户应用表',
    'm_pingxx_subapp_account':                'Ping++子商户收款账户表',
    'm_pingxx_user':                          'Ping++用户表',
    'm_trade_order':                          '交易订单主表',
    'm_trade_order_course':                   '交易订单-课程明细表',
    'm_trade_order_detail':                   '交易订单-项目明细表',
    'm_trade_order_refund':                   '交易订单退款表',
    'm_trade_settlement_bill':                '交易结算单表',
    'm_trade_settlement_bill_detail':         '交易结算单明细表',
    'm_trade_settlement_bill_detail_unionpay':'交易结算单明细表(银联)',
    'm_trade_settlement_payment_record':      '交易结算打款记录表',
    'm_trade_settlement_record':              '交易结算记录表',
    'm_trade_settlement_record_detail':       '交易结算记录明细表',
    'm_trade_stat_report':                    '交易统计报表主表',
    'm_trade_stat_report_detail':             '交易统计报表明细表',
    'm_union_pay_order':                      '银联支付订单表',
    'sys_area':                               '行政区域表',
    'sys_verification_code':                  '短信验证码表',
    't_student_jnl':                          '学生-一起吗(加能量)账号关联表',
    't_venue_sport':                          '场馆-运动项目关联表',
    't_visa_merchant':                        '银联商户信息表',
}

# ============ 通用字段中文名（无注释时兜底） ============
GENERIC_FIELD = {
    'id': '主键id',
    'gmt_create': '创建时间', 'gmt_created': '创建时间', 'create_time': '创建时间',
    'gmt_updated': '更新时间', 'update_time': '更新时间', 'up_time': '更新时间',
    'is_deleted': '是否删除', 'delete_time': '删除时间', 'delete_by': '删除人',
    'merchant_id': '商户id', 'venue_id': '场馆id(训练基地id)',
    'sport_id': '运动/课程分类id', 'role_id': '角色id', 'admin_id': '管理员id',
    'class_id': '班级id', 'teacher_id': '教师id', 'student_id': '学生id',
    'syllabus_id': '课程安排id', 'employee_id': '员工id', 'course_id': '课程id',
    'perm_code': '权限编码', 'status': '状态', 'phone': '手机号',
    'total': '总条数', 'card_no': '卡号', 'username': '用户名', 'password': '密码',
}

# ============ 针对个别无注释且非通用字段的定向补充 ============
FIELD_OVERRIDE = {
    ('m_message_tx', 'handled'):            '是否已处理',
    ('m_message_tx', 'message'):            '消息内容',
    ('m_pingxx_notify', 'notify_data_json'):'通知数据JSON',
    ('m_pingxx_notify', 'notify_event_json'):'通知事件JSON',
    ('m_pingxx_subapp', 'metadata'):        '元数据',
    ('m_trade_order_course_delay', 'opt_user_id'): '操作人id',
    ('m_trade_settlement_record', 'royalty_settle_amount'): '分润结算金额',
    ('m_trade_settlement_record_detail', 'royalty_settle_amount'): '分润结算金额',
}

def resolve_table_cn(t):
    """有注释用注释首行；无注释查推测表；返回 (中文名, 是否推测)。"""
    if t['comment']:
        # 部分注释含换行/补充说明，取首行做表名，完整注释保留信息
        cn = t['comment'].replace('\\r\\n', ' ').replace('\r\n', ' ').split('\n')[0].strip()
        return cn, False
    return TABLE_CN.get(t['name'], t['name'] + '(待确认)'), True

def resolve_field_cn(table_name, fld):
    if fld['comment']:
        return fld['comment'].strip()
    key = (table_name, fld['field'])
    if key in FIELD_OVERRIDE:
        return FIELD_OVERRIDE[key]
    return GENERIC_FIELD.get(fld['field'], fld['field'] + '(待确认)')

# ============ 生成 xlsx ============
wb = openpyxl.Workbook()

# ---- Sheet1 ----
ws1 = wb.active
ws1.title = '所有数据表及数据字段'
h1 = ['数据表名称', '数据表中文名称', '字段名称', '字段中文称', '字段类型', '截图', '别名']
ws1.append(h1)

# ---- Sheet2 ----
ws2 = wb.create_sheet('所有数据表')
h2 = ['表名', '中文表名', '引擎', '自增', '数据长度']
ws2.append(h2)

guessed = []
for t in sorted(tables, key=lambda x: x['name']):
    tcn, is_guess = resolve_table_cn(t)
    if is_guess:
        guessed.append((t['name'], tcn))
    # sheet2
    ws2.append([t['name'], tcn, t['engine'], t['auto_inc'] if t['auto_inc'] is not None else '', ''])
    # sheet1
    for fld in t['fields']:
        fcn = resolve_field_cn(t['name'], fld)
        alias = '推测表名' if is_guess else ''
        ws1.append([t['name'], tcn, fld['field'], fcn, fld['type'], '', alias])

# ============ 样式：表头加粗填充、冻结首行、列宽 ============
head_fill = PatternFill('solid', fgColor='4472C4')
head_font = Font(bold=True, color='FFFFFF')
for ws, widths in [(ws1, [30, 26, 26, 40, 22, 10, 12]), (ws2, [40, 30, 12, 12, 12])]:
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'

out = '训练_数据源清单.xlsx'
wb.save(out)
print('已生成:', out)
print('Sheet1 行数(含表头):', ws1.max_row, '| Sheet2 行数(含表头):', ws2.max_row)
print('\n本次推测中文表名的表 (%d 张):' % len(guessed))
for n, cn in guessed:
    print('   %-42s -> %s' % (n, cn))

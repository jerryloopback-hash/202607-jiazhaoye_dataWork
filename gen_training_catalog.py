# -*- coding: utf-8 -*-
"""根据 _vmdb_tables.json 生成《训练_数据源清单.xlsx》，格式严格对齐模板：
  Sheet1「所有数据表及数据字段」: 数据表名称 / 数据表中文名称 / 字段名称 / 字段中文称 / 字段类型
  Sheet2「所有数据表」:            表名 / 中文表名 / 引擎 / 自增 / 数据长度
- 有表/字段注释的：直接使用注释。
- 无注释的：根据表名、字段、同域已注释表推测中文名，并在控制台+review文件标注为"推测"。
- 数据长度：源自数据库运行态 information_schema，DDL 无此信息，留空。
"""
import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

tables = json.load(open('_vmdb_tables.json', encoding='utf-8'))

# ============================================================
# 无表注释的 111 张表 —— 人工推测中文表名
# 依据：表名语义 + 字段构成 + 同前缀已注释表的命名习惯
#   c_ 商品/进销存  d_ 数据统计  h_ 会员卡  j_ 三方/短信  l_ 平台&商户&权限
#   m_ 交易/场地/票务/租赁  o_ 大屏  p_ 停车  s_ 营收填报  sys_ 系统
#   t_ 场馆/设备/人脸/银联  theatre_ 剧场票务
# ============================================================
TABLE_CN = {
    # --- c_ 商品/进销存 ---
    'c_goods_access_storage_code':           '出入库单据编号表',
    'c_transfer_auditor_detail':             '出入库调拨审核明细表',
    # --- d_ 数据统计 ---
    'd_member_data':                         '会员数量统计表',
    'd_passenger_flow':                      '客流量统计表',
    # --- 闸机/门锁联动 ---
    'gate_link_lock':                        '闸机租柜联动记录表',
    # --- h_ 会员卡体系 ---
    'h_card':                                '实体卡表',
    'h_card_sales':                          '售卡方案表',
    'h_card_sales_detail':                   '售卡方案-可用时段明细表',
    'h_card_vip':                            '会员卡(实体卡绑定)表',
    'h_card_vip_transfer':                   '会员卡迁移记录表',
    'h_card_vip_transfer_detail':            '会员卡迁移明细表',
    'h_card_vip_transfer_detail_item':       '会员卡迁移明细-专项卡项表',
    'h_member':                              '会员信息表',
    'h_member_card':                         '会员专项卡表',
    'h_member_card_refund':                  '会员卡退款申请表',
    'h_member_charge':                       '会员充值记录表',
    'h_member_deducted':                     '会员卡抵扣记录表',
    'h_member_record':                       '会员卡消费/充值流水表',
    'h_member_record_stat_log':              '会员卡记录统计日志表',
    'h_member_record_stat_report':           '会员卡记录统计报表(汇总)表',
    'h_member_record_stat_report_detail':    '会员卡记录统计报表明细表',
    'h_member_record_stat_report_detail_2018':'会员卡记录统计报表明细表(2018年)',
    'h_member_record_stat_report_detail_2019':'会员卡记录统计报表明细表(2019年)',
    'h_member_record_stat_report_detail_2020':'会员卡记录统计报表明细表(2020年)',
    'h_member_record_stat_report_detail_2021':'会员卡记录统计报表明细表(2021年)',
    'h_member_record_stat_report_detail_2021_copy1':'会员卡记录统计报表明细表(2021年副本)',
    'h_member_record_stat_report_detail_2022':'会员卡记录统计报表明细表(2022年)',
    'h_member_record_stat_report_detail_2023':'会员卡记录统计报表明细表(2023年)',
    'h_member_record_stat_report_detail_2024':'会员卡记录统计报表明细表(2024年)',
    'h_member_record_stat_report_detail_2025':'会员卡记录统计报表明细表(2025年)',
    'h_member_record_stat_report_detail_2026':'会员卡记录统计报表明细表(2026年)',
    'h_member_record_statistic':             '会员卡记录统计表',
    'h_member_record_statistic_100':         '会员卡记录统计表(分表100)',
    'h_member_validcode':                    '会员手机验证码表',
    'h_random_visitor':                      '随机访客数按日统计表',
    # --- j_ 三方/文件/短信 ---
    'j_upload_file_return':                  '上传文件回盘记录表',
    # --- l_ 平台&商户&权限 ---
    'l_admin':                               '平台管理员表',
    'l_field_board_line':                    '场地面板-线路关联表',
    'l_merchant_settle_account':             '商户结算收款账户表',
    'l_permission':                          '平台权限资源表',
    'l_sport':                               '运动类型表',
    # --- m_ 交易/场地/票务/租赁 ---
    'm_coupon':                              '商户优惠券表',
    'm_field_fixed':                         '固定场/免费场锁定表',
    'm_field_fixed_detail':                  '固定场时段明细表',
    'm_field_fixed_enter_record':            '固定场入场核销记录表',
    'm_field_fixed_free':                    '免费场订场条件表',
    'm_field_fixed_inner':                   '固定场-场地关联表',
    'm_field_fixed_week':                    '固定场-周几关联表',
    'm_field_lock_type':                     '锁场类目表',
    'm_field_price_inner':                   '场地价格-场地关联表',
    'm_field_record_detail':                 '场地操作记录明细表',
    'm_group_ticket_give':                   '团体票赠送表',
    'm_group_ticket_give_no':                '团体票赠送票号表',
    'm_group_ticket_give_record':            '团体票赠送领取记录表',
    'm_group_ticket_online_sell':            '团体票线上销售设置表',
    'm_inform':                              '通知消息表',
    'm_leased_goods':                        '租赁商品表',
    'm_leased_goods_category':               '租赁商品分类表',
    'm_leased_goods_stocktaking':            '租赁商品盘点记录表',
    'm_leased_timeunit':                     '租赁时间单位表',
    'm_sys_type_value':                      '租户端类型字典表',
    'm_trade_order':                         '交易订单主表',
    'm_trade_order_coupon':                  '交易订单-优惠券使用表',
    'm_trade_order_detail':                  '交易订单明细表',
    'm_trade_order_field':                   '交易订单-场地预定表',
    'm_trade_order_field_gate':              '交易订单-场地入闸核销表',
    'm_trade_order_field_gate_record':       '交易订单-场地入闸核销记录表',
    'm_trade_order_leased':                  '交易订单-租赁订单表',
    'm_trade_order_leased_goods':            '交易订单-租赁商品明细表',
    'm_trade_order_pay':                     '交易订单支付表',
    'm_trade_order_pay_param':               '交易订单支付参数表',
    'm_trade_order_refund':                  '交易订单退款表',
    'm_trade_order_refund_offline':          '交易订单线下退款表',
    'm_trade_order_type':                    '交易订单类型字典表',
    'm_trade_settlement_bill_unionpay':      '交易结算单表(银联)',
    'm_trade_stat_report_detail':            '交易统计报表明细表',
    'm_union_id_b_link':                     '通票会员-B端用户id关联表',
    'm_union_pay_order':                     '银联支付订单表',
    'm_union_pay_order_abc':                 '农行支付订单表',
    'm_union_ticket':                        '通票表',
    'm_union_ticket_detail':                 '通票票号明细表',
    'm_venue_customer':                      '场馆顾客表',
    # --- o_ 大屏 ---
    'o_screen_fitness_activity':             '大屏全民健身活动数据表',
    'o_screen_setting':                      '大屏设置表',
    # --- p_ 停车 ---
    'p_park_jieshun_config':                 '停车场捷顺对接配置表',
    'p_park_record':                         '停车记录表',
    'p_park_ticket':                         '停车券表',
    'p_park_ticket_code':                    '停车券方案表',
    'p_park_ticket_code_detail':             '停车券方案券码明细表',
    'p_park_ticket_config':                  '停车券发放配置表',
    'p_park_ticket_config_detail':           '停车券发放配置明细表',
    # --- sys_ 系统 ---
    'sys_area':                              '行政区域表',
    # --- t_ 场馆/设备/人脸/银联/门锁 ---
    't_door_lock_admin_code':                '门锁管理员二维码表',
    't_door_lock_enter_record':              '门锁进门记录表',
    't_face_capture':                        '人脸抓拍记录表',
    't_face_flow':                           '人脸客流统计表',
    't_face_flow_group':                     '人脸客流统计组表',
    't_face_organization':                   '海康人脸组织表',
    't_face_person':                         '海康人脸人员表',
    't_face_push':                           '人脸推送记录表',
    't_gate_enter_time':                     '闸机可入场时段表',
    't_hk_resource':                         '海康设备资源表',
    't_venue_config':                        '场馆检票/识别配置表',
    't_visa_merchant':                       '银联商户信息表',
    't_visa_merchant_abc':                   '农行二级商户信息表',
    't_visa_third_config':                   '银联第三方分润配置表',
    't_xly_token':                           '小厘云授权Token表',
    # --- theatre_ 剧场票务 ---
    'theatre_ticket_discount':               '剧场票务-座位区折扣表',
    'theatre_ticket_item':                   '剧场票务-演出项目表',
    'theatre_ticket_price':                  '剧场票务-座位区票价表',
    'theatre_ticket_third_order':            '剧场票务-第三方订单表',
}

# ============================================================
# 通用字段中文名（无字段注释时兜底）
# ============================================================
GENERIC_FIELD = {
    'id': '主键id',
    'gmt_create': '创建时间', 'gmt_created': '创建时间', 'gmt_create_time': '创建时间',
    'create_time': '创建时间', 'creat_time': '创建时间', 'created_time': '创建时间',
    'gmt_updated': '更新时间', 'gmt_update_time': '更新时间', 'update_time': '更新时间',
    'updated_time': '更新时间', 'up_time': '更新时间',
    'is_deleted': '是否删除', 'is_delete': '是否删除',
    'delete_time': '删除时间', 'delete_by': '删除人',
    'create_user': '创建人', 'update_user': '更新人',
    'create_user_id': '创建人id', 'update_user_id': '更新人id',
    'create_name': '创建人名称', 'create_user_name': '创建人名称',
    'sort_num': '排序号',
    'merchant_id': '商户id', 'venue_id': '场馆id',
    'sport_id': '运动类型id', 'role_id': '角色id', 'admin_id': '管理员id',
    'employee_id': '员工id', 'organization_id': '机构id',
    'perm_code': '权限编码', 'status': '状态', 'phone': '手机号',
    'card_no': '卡号', 'username': '用户名', 'password': '密码',
    'business_id': '业务id', 'tag_id': '标签id',
    # 会员卡统计报表 h_member_record_stat_report[_detail][_YYYY] 专用字段
    'card_sale_id': '售卡id', 'card_name': '卡名称', 'card_type': '卡类型',
    'member_card_id': '专项卡id',
    'start_amount': '期初金额', 'start_num': '期初次数',
    'charge_amount': '充值金额', 'charge_num': '充值次数',
    'consume_amount': '消费金额', 'consume_num': '消费次数',
    'end_amount': '期末金额', 'end_num': '期末次数',
    'as_revenue_amount': '确认收入金额', 'as_revenue_num': '确认收入次数',
    'as_revenue_expire_amount': '确认收入金额(卡过期)', 'as_revenue_expire_num': '确认收入次数(卡过期)',
    'as_revenue_disable_amount': '确认收入金额(卡作废)', 'as_revenue_disable_num': '确认收入次数(卡作废)',
    'stat_time': '统计时间',
}

# ============================================================
# 定向字段中文名（同名字段在不同表含义不同 / 无注释且非通用）
# key = (表名, 字段名)
# ============================================================
FIELD_OVERRIDE = {
    ('c_goods', 'is_deleted'): '是否删除',
    ('c_goods_daily_report', 'receiver_code'): '接收方批次',
    ('c_goods_barcode', 'venue_id'): '场馆id',
    ('c_goods_market', 'market_id'): '商超id',
    ('c_goods_receiver_code', 'access_storage_id'): '出入库单据id',
    ('c_goods_receiver_code', 'in_num'): '入库数量',
    ('c_goods_receiver_code', 'price'): '单价',
    ('c_goods_receiver_code', 'tax_rate'): '税率',
    ('c_goods_receiver_code', 'receiver_code'): '接收方批次',
    ('c_transfer_auditor_detail', 'auditor_user_id'): '审核人id',
    ('gate_link_lock', 'gate_open_time'): '开闸时间',
    ('gate_link_lock', 'gate_close_time'): '出闸时间',
    ('gate_link_lock', 'lock_open_time'): '开柜时间',
    ('gate_link_lock', 'lock_close_time'): '还柜时间',
    ('gate_link_lock', 'gate_close_from'): '出闸方式',
    ('h_card_sales', 'online_sell_stock'): '线上销售库存',
    ('h_card_vip_transfer_detail_item', 'transfer_detail_id'): '迁移明细id',
    ('h_member_card', 'prepaid_balance'): '预收入余额',
    ('h_member_card_refund', 'item_id'): '项目id',
    ('h_member_card_refund', 'item_name'): '项目名称',
    ('h_member_delay_record', 'op_user_name'): '操作人名称',
    ('h_member_record_stat_log', 'record_id'): '记录id',
    ('h_member_record_stat_log', 'success'): '是否成功',
    ('h_member_record_stat_log', 'content'): '内容',
    ('j_app_version', 'update_user'): '更新人',
    ('j_app_version', 'create_user'): '创建人',
    ('j_member_wx_palm', 'phone'): '手机号',
    ('j_sms', 'phone'): '手机号',
    ('j_sms_template', 'sms_key'): '短信模板key',
    ('l_admin_role', 'admin_id'): '管理员id',
    ('l_lamp_account', 'id'): '主键id',
    ('m_card_event', 'guid'): '全局唯一id',
    ('m_card_event', 'message'): '消息内容',
    ('m_enter_gate', 'gmt_create'): '创建时间',
    ('m_escort_card', 'enter_time'): '入场时间',
    ('m_gate_event', 'message'): '消息内容',
    ('m_gate_event', 'gate_timestamp'): '闸机时间戳',
    ('m_group_ticket_give', 'venue_name'): '场馆名称',
    ('m_group_ticket_give_record', 'give_id'): '赠送id',
    ('m_inform', 'op_date'): '操作时间',
    ('m_inform_read', 'employee_id'): '员工id',
    ('m_inform_read', 'inform_id'): '通知id',
    ('m_message_tx', 'handled'): '是否已处理',
    ('m_message_tx', 'message'): '消息内容',
    ('m_sys_employee_role', 'employee_id'): '员工id',
    ('m_sys_revenue_permission', 'employee_id'): '员工id',
    ('m_sys_role_permission', 'role_id'): '角色id',
    ('m_trade_order_field_gate', 'field_id'): '场地id',
    ('m_trade_order_locker_rent', 'venue_id'): '场馆id',
    ('m_trade_order_pay_param', 'pay_params'): '支付参数',
    ('m_trade_order_pay_param', 'create_order_param'): '下单参数',
    ('m_trade_order_ticket_no', 'venue_id'): '场馆id',
    ('m_trade_settlement_application', 'created_by'): '创建人',
    ('m_trade_settlement_application', 'created_time'): '创建时间',
    ('m_union_id_b_link', 'm_union_id'): '会员端通票用户id',
    ('m_union_id_b_link', 'b_union_id'): 'B端通票用户id',
    ('o_oa_message', 'operator_id'): '操作人id',
    ('s_revenue_budget_amount', 'budget_id'): '预算id',
    ('s_revenue_budget_amount', 'organization_id'): '机构id',
    ('s_revenue_collect', 'organization_id'): '机构id',
    ('s_revenue_collect', 'employee_id'): '员工id',
    ('s_revenue_collect_amount', 'collect_id'): '填报id',
    ('s_revenue_template', 'organization_id'): '机构id',
    ('s_revenue_template_employee', 'employee_id'): '员工id',
    ('s_revenue_template_employee', 'organization_id'): '机构id',
    ('s_revenue_template_type', 'template_id'): '模板id',
    ('t_face_push', 'update_time'): '更新时间',
    ('t_gate_enter_time', 'start_time'): '开始时间',
    ('t_gate_enter_time', 'end_time'): '结束时间',
    ('t_hk_resource', 'jzy_resource_id'): '佳兆业设备资源id',
    ('t_organization_employee', 'organization_id'): '机构id',
    ('t_organization_employee', 'employee_id'): '员工id',
    ('t_venue_business', 'business_id'): '业务id',
    ('t_venue_employee', 'employee_id'): '员工id',
    ('t_venue_tag_ref', 'tag_id'): '标签id',
    ('t_venue_tag_ref_snapshot', 'tag_id'): '标签id',
    ('theatre_sale_channel_detail', 'channel_id'): '分销渠道id',
    ('theatre_sale_channel_detail', 'channel_source'): '渠道来源',
    ('theatre_sale_channel_detail', 'channel_com_name'): '渠道公司名称',
    ('theatre_ticket_item', 'third_project_id'): '第三方项目id',
    ('theatre_ticket_item', 'third_project_type'): '第三方项目类型',
    ('theatre_ticket_lock_log', 'sell_detail_ids'): '售票明细id列表',
    ('theatre_ticket_lock_log', 'operator'): '操作人',
    ('theatre_ticket_log', 'content'): '内容',
    ('theatre_ticket_log', 'result_success'): '是否成功',
    ('theatre_ticket_log', 'error_message'): '错误信息',
    ('theatre_ticket_log', 'operator'): '操作人',
    ('theatre_ticket_price', 'colour'): '颜色',
    ('theatre_ticket_price', 'name'): '名称',
    ('theatre_ticket_price', 'plan_id'): '方案id',
    ('theatre_ticket_price', 'price'): '价格',
    ('theatre_ticket_seat', 'seat_id'): '座位id',
    ('theatre_ticket_seat_area', 'code'): '编码',
    ('theatre_ticket_seat_area', 'name'): '名称',
    ('theatre_ticket_seat_area', 'plan_name'): '方案名称',
    ('theatre_ticket_seat_area', 'plan_id'): '方案id',
    ('theatre_ticket_seat_area', 'price'): '价格',
    ('theatre_ticket_seat_area', 'create_name'): '创建人名称',
    ('theatre_ticket_sell', 'item_id'): '演出项目id',
    ('theatre_ticket_sell', 'sale_start_time'): '售卖开始时间',
    ('theatre_ticket_sell', 'sale_end_time'): '售卖结束时间',
    ('theatre_ticket_sell', 'third_project_id'): '第三方项目id',
    ('theatre_ticket_sell', 'third_project_type'): '第三方项目类型',
    ('theatre_ticket_sell_detail', 'seat_num_x'): '座位坐标X',
    ('theatre_ticket_sell_detail', 'seat_num_y'): '座位坐标Y',
    ('theatre_ticket_sell_detail', 'seat_row_no'): '座位排号',
    ('theatre_ticket_sell_detail', 'seat_col_no'): '座位列号',
    ('theatre_ticket_sell_detail', 'is_side'): '是否边座',
    ('theatre_ticket_sell_detail', 'price_id'): '票价id',
    ('theatre_ticket_sell_detail', 'channel_com_name'): '渠道公司名称',
    ('theatre_ticket_third_order', 'third_order_num'): '第三方订单号',
    ('theatre_ticket_third_order', 'third_type'): '第三方类型',
    ('theatre_ticket_third_order', 'order_time'): '下单时间',
    ('theatre_ticket_third_order', 'total_amount'): '订单总金额',
    ('theatre_ticket_third_order', 'real_amount'): '实付金额',
    ('theatre_sale_channel', 'is_deleted'): '是否删除',
    ('theatre_ticket_discount', 'create_time'): '创建时间',
    ('ticket_revenue_stat', 'venue_name'): '场馆名称',
    ('ticket_revenue_stat', 'ticket_id'): '票id',
    ('ticket_revenue_stat', 'ticket_no_id'): '票号id',
    ('ticket_revenue_stat', 'ticket_no'): '票号',
    ('ticket_revenue_stat_report', 'venue_name'): '场馆名称',
    ('ticket_revenue_stat_report', 'ticket_id'): '票id',
}

PLACEHOLDER = '(推测,待确认)'

def resolve_table_cn(t):
    """返回 (中文名, 是否推测)。有注释取首行；无注释查推测表。"""
    if t['comment']:
        cn = t['comment'].replace('\\r\\n', ' ').replace('\r\n', ' ').replace('\r', ' ')
        cn = cn.split('\n')[0].strip()
        return cn, False
    if t['name'] in TABLE_CN:
        return TABLE_CN[t['name']], True
    return t['name'] + PLACEHOLDER, True

def resolve_field_cn(table_name, fld):
    """返回 (中文名, 是否兜底占位)。"""
    if fld['comment']:
        return fld['comment'].strip(), False
    key = (table_name, fld['field'])
    if key in FIELD_OVERRIDE:
        return FIELD_OVERRIDE[key], False
    if fld['field'] in GENERIC_FIELD:
        return GENERIC_FIELD[fld['field']], False
    return fld['field'] + PLACEHOLDER, True

# ============================================================
# 生成 xlsx
# ============================================================
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = '所有数据表及数据字段'
ws1.append(['数据表名称', '数据表中文名称', '字段名称', '字段中文称', '字段类型'])

ws2 = wb.create_sheet('所有数据表')
ws2.append(['表名', '中文表名', '引擎', '自增', '数据长度'])

guessed_tables = []
placeholder_fields = []
for t in sorted(tables, key=lambda x: x['name']):
    tcn, is_guess = resolve_table_cn(t)
    if is_guess:
        guessed_tables.append((t['name'], tcn))
    ws2.append([t['name'], tcn, t['engine'],
                t['auto_inc'] if t['auto_inc'] is not None else '', ''])
    for fld in t['fields']:
        fcn, is_ph = resolve_field_cn(t['name'], fld)
        if is_ph:
            placeholder_fields.append((t['name'], fld['field']))
        ws1.append([t['name'], tcn, fld['field'], fcn, fld['type']])

# 样式：表头加粗蓝底白字、冻结首行、列宽
head_fill = PatternFill('solid', fgColor='4472C4')
head_font = Font(bold=True, color='FFFFFF')
for ws, widths in [(ws1, [34, 30, 28, 46, 22]), (ws2, [40, 34, 12, 12, 12])]:
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'

import os
out = '训练_数据源清单.xlsx'
try:
    wb.save(out)
except PermissionError:
    out = '训练_数据源清单_new.xlsx'
    wb.save(out)
    print('[提示] 原文件被占用(可能在Excel中打开)，已改存为:', out)
print('已生成:', out)
print('Sheet1「所有数据表及数据字段」行数(含表头):', ws1.max_row)
print('Sheet2「所有数据表」行数(含表头):', ws2.max_row)
print('\n推测中文表名的表: %d 张' % len(guessed_tables))
print('字段仍用占位(待确认)的: %d 个' % len(placeholder_fields))

# 输出推测清单到 review 文件，便于人工复核
with open('_推测表名复核清单.txt', 'w', encoding='utf-8') as f:
    f.write('# 推测中文表名清单（共 %d 张，需人工复核）\n\n' % len(guessed_tables))
    for n, cn in guessed_tables:
        f.write('%-46s -> %s\n' % (n, cn))
    f.write('\n\n# 仍用字段名占位的字段（共 %d 个）\n\n' % len(placeholder_fields))
    for tn, fn in placeholder_fields:
        f.write('%-46s . %s\n' % (tn, fn))
print('推测复核清单已写出: _推测表名复核清单.txt')
